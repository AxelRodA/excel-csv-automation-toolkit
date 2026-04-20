import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "sample-data" / "company_leads_dirty.csv"
XLSX_PATH = ROOT / "sample-data" / "company_leads_dirty.xlsx"


def main() -> None:
    rows = list(csv.reader(CSV_PATH.read_text(encoding="utf-8").splitlines()))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dirty Leads"

    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    table_ref = f"A1:F{sheet.max_row}"
    table = Table(displayName="DirtyCompanyLeads", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"

    widths = {"A": 26, "B": 30, "C": 20, "D": 14, "E": 18, "F": 16}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(XLSX_PATH)
    print(f"Created {XLSX_PATH}")


if __name__ == "__main__":
    main()
